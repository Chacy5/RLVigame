import os
import logging
import random
import asyncio
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional
from urllib.parse import urlparse

import pg8000
from openpyxl import load_workbook
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import (
    ApplicationBuilder,
    ContextTypes,
    CommandHandler,
    CallbackQueryHandler,
)

# ----------------- Логирование -----------------
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ----------------- Переменные окружения -----------------
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
DATABASE_URL = os.getenv("DATABASE_URL")
PARTNER_CHAT_ID = os.getenv("PARTNER_CHAT_ID")  # опционально, chat_id партнёра

if not TELEGRAM_BOT_TOKEN:
    logger.error("TELEGRAM_BOT_TOKEN is not set")
    raise SystemExit("TELEGRAM_BOT_TOKEN is required")

if not DATABASE_URL:
    logger.error("DATABASE_URL is not set")
    raise SystemExit("DATABASE_URL is required")

# ----------------- Конфиг БД -----------------

DB_CONFIG: Dict[str, object] = {}


def parse_db_url(url: str) -> None:
    global DB_CONFIG
    parsed = urlparse(url)
    DB_CONFIG = {
        "user": parsed.username,
        "password": parsed.password,
        "host": parsed.hostname,
        "port": parsed.port or 5432,
        "database": parsed.path.lstrip("/"),
    }
    logger.info("DB config parsed: host=%s db=%s", DB_CONFIG["host"], DB_CONFIG["database"])


def get_connection():
    return pg8000.connect(**DB_CONFIG)


# ----------------- Структуры данных -----------------


@dataclass
class LootBoxReward:
    roll: int
    text: str


LOOTBOX_TABLES: Dict[int, List[LootBoxReward]] = {}
MINI_EVENTS: List[str] = []

# триггеры уведомлений партнёру
PARTNER_KEYWORDS = [
    "от него",
    "от парня",
    "свидание",
    "завтрак в постель",
    "кофе в постель",
    "массаж",
    "обнимаш",
]


def load_lootboxes_from_excel(path: str = "Лутбоксы.xlsx") -> None:
    """
    Загружаем 5 d100-таблиц лутбоксов + мини-ивенты из Excel.
    """
    global LOOTBOX_TABLES, MINI_EVENTS
    logger.info("Загружаю лутбоксы и мини-ивенты из '%s'...", path)
    wb = load_workbook(path, data_only=True)

    # Первые 5 листов — лутбоксы (1..5)
    for idx, box_number in enumerate(range(1, 6), start=0):
        ws = wb[wb.sheetnames[idx]]
        rewards: List[LootBoxReward] = []
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            key = row[0]
            if isinstance(key, (int, float)) and row[1]:
                roll = int(key)
                text = str(row[1]).strip()
                rewards.append(LootBoxReward(roll=roll, text=text))
        rewards.sort(key=lambda r: r.roll)
        LOOTBOX_TABLES[box_number] = rewards

    # Последний лист — мини-ивенты
    ws_me = wb[wb.sheetnames[-1]]
    rows = [r[0] for r in ws_me.iter_rows(values_only=True)]

    import re

    events: List[str] = []
    current_lines: List[str] = []
    for cell in rows:
        if not cell:
            continue
        text = str(cell).strip()
        # строки вида "1. ..." – начало нового ивента
        if re.match(r"^\D*\d+\.", text):
            if current_lines:
                events.append("\n".join(current_lines))
                current_lines = []
            current_lines.append(text)
        else:
            if current_lines:
                current_lines.append(text)
    if current_lines:
        events.append("\n".join(current_lines))

    MINI_EVENTS = events
    logger.info(
        "Лутбоксы и мини-ивенты загружены. Box1=%s, Box2=%s, Box3=%s, Box4=%s, Box5=%s, mini_events=%s",
        len(LOOTBOX_TABLES.get(1, [])),
        len(LOOTBOX_TABLES.get(2, [])),
        len(LOOTBOX_TABLES.get(3, [])),
        len(LOOTBOX_TABLES.get(4, [])),
        len(LOOTBOX_TABLES.get(5, [])),
        len(MINI_EVENTS),
    )


# ----------------- Работа с БД (pg8000, через to_thread) -----------------


async def init_db() -> None:
    def _work():
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                user_id     BIGINT PRIMARY KEY,
                username    TEXT,
                first_name  TEXT,
                coins       INTEGER NOT NULL DEFAULT 0,
                created_at  TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS reward_cards (
                id          SERIAL PRIMARY KEY,
                user_id     BIGINT NOT NULL REFERENCES users(user_id) ON DELETE CASCADE,
                box_type    SMALLINT NOT NULL,
                roll        INTEGER NOT NULL,
                reward_text TEXT NOT NULL,
                is_opened   BOOLEAN NOT NULL DEFAULT FALSE,
                is_used     BOOLEAN NOT NULL DEFAULT FALSE,
                created_at  TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS inventory_items (
                id              SERIAL PRIMARY KEY,
                user_id         BIGINT NOT NULL REFERENCES users(user_id) ON DELETE CASCADE,
                description     TEXT NOT NULL,
                source_card_id  INTEGER REFERENCES reward_cards(id),
                is_consumed     BOOLEAN NOT NULL DEFAULT FALSE,
                created_at      TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
        conn.commit()
        cur.close()
        conn.close()

    logger.info("Инициализирую схему БД...")
    await asyncio.to_thread(_work)
    logger.info("Схема БД инициализирована.")


async def get_or_create_user(user_id: int, username: str, first_name: str) -> Dict:
    def _work():
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT user_id, username, first_name, coins FROM users WHERE user_id=%s",
            (user_id,),
        )
        row = cur.fetchone()
        if not row:
            cur.execute(
                "INSERT INTO users(user_id, username, first_name) VALUES(%s,%s,%s)",
                (user_id, username, first_name),
            )
            conn.commit()
            cur.execute(
                "SELECT user_id, username, first_name, coins FROM users WHERE user_id=%s",
                (user_id,),
            )
            row = cur.fetchone()
        cur.close()
        conn.close()
        return {"user_id": row[0], "username": row[1], "first_name": row[2], "coins": row[3]}

    return await asyncio.to_thread(_work)


async def get_user(user_id: int) -> Optional[Dict]:
    def _work():
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT user_id, username, first_name, coins FROM users WHERE user_id=%s",
            (user_id,),
        )
        row = cur.fetchone()
        cur.close()
        conn.close()
        if not row:
            return None
        return {"user_id": row[0], "username": row[1], "first_name": row[2], "coins": row[3]}

    return await asyncio.to_thread(_work)


async def update_coins(user_id: int, delta: int) -> int:
    def _work():
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            "UPDATE users SET coins = coins + %s WHERE user_id=%s RETURNING coins",
            (delta, user_id),
        )
        row = cur.fetchone()
        conn.commit()
        cur.close()
        conn.close()
        return row[0]

    return await asyncio.to_thread(_work)


async def add_reward_card(
    user_id: int, box_type: int, roll: int, reward_text: str
) -> int:
    def _work():
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO reward_cards(user_id, box_type, roll, reward_text)
            VALUES(%s,%s,%s,%s)
            RETURNING id
            """,
            (user_id, box_type, roll, reward_text),
        )
        row = cur.fetchone()
        conn.commit()
        cur.close()
        conn.close()
        return row[0]

    return await asyncio.to_thread(_work)


async def list_reward_cards(user_id: int) -> List[Dict]:
    def _work():
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            """
            SELECT id, box_type, roll, reward_text, is_opened, is_used, created_at
            FROM reward_cards
            WHERE user_id=%s
            ORDER BY created_at DESC
            """,
            (user_id,),
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()
        res = []
        for r in rows:
            res.append(
                {
                    "id": r[0],
                    "box_type": r[1],
                    "roll": r[2],
                    "reward_text": r[3],
                    "is_opened": r[4],
                    "is_used": r[5],
                    "created_at": r[6],
                }
            )
        return res

    return await asyncio.to_thread(_work)


async def open_reward_card(card_id: int) -> Optional[Dict]:
    def _work():
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            """
            UPDATE reward_cards
            SET is_opened = TRUE
            WHERE id=%s AND is_opened = FALSE
            RETURNING id, user_id, box_type, roll, reward_text, is_opened, is_used, created_at
            """,
            (card_id,),
        )
        row = cur.fetchone()
        if not row:
            conn.commit()
            cur.close()
            conn.close()
            return None
        # сразу кладём в инвентарь
        cur.execute(
            """
            INSERT INTO inventory_items(user_id, description, source_card_id)
            VALUES(%s,%s,%s)
            """,
            (row[1], row[4], row[0]),
        )
        conn.commit()
        cur.close()
        conn.close()
        return {
            "id": row[0],
            "user_id": row[1],
            "box_type": row[2],
            "roll": row[3],
            "reward_text": row[4],
            "is_opened": row[5],
            "is_used": row[6],
            "created_at": row[7],
        }

    return await asyncio.to_thread(_work)


async def list_inventory(user_id: int) -> List[Dict]:
    def _work():
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            """
            SELECT id, description, is_consumed, created_at
            FROM inventory_items
            WHERE user_id=%s
            ORDER BY created_at DESC
            """,
            (user_id,),
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()
        res = []
        for r in rows:
            res.append(
                {
                    "id": r[0],
                    "description": r[1],
                    "is_consumed": r[2],
                    "created_at": r[3],
                }
            )
        return res

    return await asyncio.to_thread(_work)


async def consume_inventory_item(item_id: int) -> Optional[Dict]:
    def _work():
        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            """
            UPDATE inventory_items
            SET is_consumed = TRUE
            WHERE id=%s AND is_consumed = FALSE
            RETURNING id, user_id, description, is_consumed, created_at
            """,
            (item_id,),
        )
        row = cur.fetchone()
        conn.commit()
        cur.close()
        conn.close()
        if not row:
            return None
        return {
            "id": row[0],
            "user_id": row[1],
            "description": row[2],
            "is_consumed": row[3],
            "created_at": row[4],
        }

    return await asyncio.to_thread(_work)


async def reset_user(user_id: int) -> None:
    def _work():
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM users WHERE user_id=%s", (user_id,))
        conn.commit()
        cur.close()
        conn.close()

    await asyncio.to_thread(_work)


# ----------------- Логика игры -----------------


def reward_for_box(box_type: int) -> Tuple[int, str]:
    rewards = LOOTBOX_TABLES.get(box_type)
    if not rewards:
        raise ValueError(f"Unknown lootbox type {box_type}")
    roll = random.randint(1, 100)
    idx = min(max(roll - 1, 0), len(rewards) - 1)
    text = rewards[idx].text
    return roll, text


def partner_should_be_notified(reward_text: str) -> bool:
    lower = reward_text.lower()
    return any(k in lower for k in PARTNER_KEYWORDS)


# ----------------- Хендлеры -----------------


async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    await get_or_create_user(
        user.id, user.username or "", user.first_name or user.full_name or ""
    )
    text = (
        "Привет! Это твоя личная RPG-игра для жизни.\n\n"
        "• Выполняй квесты и получай монеты.\n"
        "• Покупай лутбоксы и открывай карты-награды.\n"
        "• Все полученные награды складываются в инвентарь.\n\n"
        "Команды:\n"
        "/profile – твой профиль\n"
        "/openbox – купить и открыть лутбокс\n"
        "/cards – все твои карты-награды\n"
        "/inventory – инвентарь\n"
        "/mini – вытянуть мини-ивент дня\n"
        "/reset – сбросить игру (осторожно)"
    )
    await update.message.reply_text(text)


async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await cmd_start(update, context)


async def cmd_profile(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    row = await get_user(user.id)
    if not row:
        row = await get_or_create_user(
            user.id, user.username or "", user.first_name or user.full_name or ""
        )

    # Для красоты можно посчитать кол-во карт и предметов
    async def _counts():
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM reward_cards WHERE user_id=%s", (user.id,))
        cards_count = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM inventory_items WHERE user_id=%s", (user.id,))
        inv_count = cur.fetchone()[0]
        cur.close()
        conn.close()
        return cards_count, inv_count

    cards_count, inv_count = await asyncio.to_thread(_counts)

    text = (
        f"Профиль {user.first_name}:\n"
        f"Монеты: {row['coins']}\n"
        f"Карт наград: {cards_count}\n"
        f"Предметов в инвентаре: {inv_count}"
    )
    await update.message.reply_text(text)


async def cmd_add_coins(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    if not context.args:
        await update.message.reply_text("Использование: /addcoins <число>")
        return
    try:
        delta = int(context.args[0])
    except ValueError:
        await update.message.reply_text("Нужно число, например /addcoins 10")
        return
    new_balance = await update_coins(user.id, delta)
    await update.message.reply_text(
        f"Ты получила {delta} монет. Новый баланс: {new_balance}"
    )


async def cmd_openbox(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    keyboard = [
        [
            InlineKeyboardButton(
                "1. Little Happiness (10)", callback_data="buy_box:1"
            ),
            InlineKeyboardButton("2. Средний (20)", callback_data="buy_box:2"),
        ],
        [
            InlineKeyboardButton("3. Большой (40)", callback_data="buy_box:3"),
        ],
        [
            InlineKeyboardButton("4. Эпический (80)", callback_data="buy_box:4"),
            InlineKeyboardButton("5. Легендарный (150)", callback_data="buy_box:5"),
        ],
    ]
    await update.message.reply_text(
        "Выбери лутбокс для покупки.\nСтоимость в монетах:",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )


BOX_COSTS = {1: 10, 2: 20, 3: 40, 4: 80, 5: 150}


async def cb_buy_box(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    _, box_str = query.data.split(":")
    box_type = int(box_str)
    user = query.from_user

    row = await get_user(user.id)
    if not row:
        row = await get_or_create_user(
            user.id, user.username or "", user.first_name or user.full_name or ""
        )
    balance = row["coins"]
    cost = BOX_COSTS[box_type]

    if balance < cost:
        await query.edit_message_text(
            f"Не хватает монет. Нужно {cost}, у тебя сейчас {balance}."
        )
        return

    await update_coins(user.id, -cost)
    roll, reward_text = reward_for_box(box_type)
    card_id = await add_reward_card(user.id, box_type, roll, reward_text)

    msg = (
        f"✨ Ты купила лутбокс {box_type} за {cost} монет.\n"
        "Бросаем кость d100...\n"
        f"Выпало: {roll}.\n\n"
        "🃏 Тебе выпала карта-награда, но она пока закрыта.\n"
        "Нажми кнопку ниже, чтобы открыть её!"
    )
    keyboard = [
        [InlineKeyboardButton("Открыть карту 🎴", callback_data=f"open_card:{card_id}")]
    ]
    await query.edit_message_text(msg, reply_markup=InlineKeyboardMarkup(keyboard))

    if PARTNER_CHAT_ID and partner_should_be_notified(reward_text):
        try:
            await context.bot.send_message(
                chat_id=int(PARTNER_CHAT_ID),
                text=(
                    f"🔔 {user.first_name} вытянула награду, которая связана с тобой:\n"
                    f"«{reward_text}».\n"
                    "Ты знаешь, что делать 😉"
                ),
            )
        except Exception as e:
            logger.warning("Не удалось отправить сообщение партнёру: %s", e)


async def cb_open_card(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    _, card_str = query.data.split(":")
    card_id = int(card_str)
    row = await open_reward_card(card_id)
    if not row:
        await query.edit_message_text("Карта не найдена или уже открыта.")
        return
    text = (
        "🎴 Карта раскрыта!\n\n"
        f"Награда:\n{row['reward_text']}\n\n"
        "Она добавлена в твой инвентарь. Когда используешь — отметь это через /inventory."
    )
    await query.edit_message_text(text)


async def cmd_cards(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    cards = await list_reward_cards(user.id)
    if not cards:
        await update.message.reply_text("У тебя пока нет карт-наград.")
        return
    lines = []
    for c in cards[:30]:
        status = "✅ открыта" if c["is_opened"] else "🔒 закрыта"
        lines.append(f"#{c['id']} · Лутбокс {c['box_type']} · {status}")
        if c["is_opened"]:
            lines.append(f"    {c['reward_text']}")
    await update.message.reply_text("Твои карты-награды:\n\n" + "\n".join(lines))


async def cmd_inventory(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    items = await list_inventory(user.id)
    if not items:
        await update.message.reply_text("Инвентарь пуст.")
        return
    lines = []
    keyboard_rows = []
    for it in items[:40]:
        status = " (использовано)" if it["is_consumed"] else ""
        lines.append(f"#{it['id']}{status}: {it['description']}")
        if not it["is_consumed"]:
            keyboard_rows.append(
                [
                    InlineKeyboardButton(
                        f"Использовать #{it['id']}",
                        callback_data=f"use_item:{it['id']}",
                    )
                ]
            )
    text = "🎒 Инвентарь:\n\n" + "\n".join(lines)
    reply_markup = InlineKeyboardMarkup(keyboard_rows) if keyboard_rows else None
    await update.message.reply_text(text, reply_markup=reply_markup)


async def cb_use_item(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    _, item_str = query.data.split(":")
    item_id = int(item_str)
    row = await consume_inventory_item(item_id)
    if not row:
        await query.edit_message_text("Этот предмет уже был использован или не найден.")
        return
    text = (
        f"Ты пометила награду как использованную:\n\n{row['description']}\n\n"
        "Хорошая работа ✨"
    )
    await query.edit_message_text(text)


async def cmd_mini(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not MINI_EVENTS:
        await update.message.reply_text("Мини-ивенты не найдены в файле Excel.")
        return
    event = random.choice(MINI_EVENTS)
    text = "🎲 Тянем мини-ивент дня...\n\n" + event
    await update.message.reply_text(text)


async def cmd_reset(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    await reset_user(user.id)
    await update.message.reply_text(
        "Твой прогресс полностью сброшен. Можно начинать заново с /start."
    )


# ----------------- Старт приложения -----------------


async def on_startup(app):
    parse_db_url(DATABASE_URL)
    load_lootboxes_from_excel("Лутбоксы.xlsx")
    await init_db()


def main() -> None:
    logger.info("Запуск бота (режим long polling)...")

    application = (
        ApplicationBuilder()
        .token(TELEGRAM_BOT_TOKEN)
        .post_init(on_startup)
        .build()
    )

    # команды
    application.add_handler(CommandHandler("start", cmd_start))
    application.add_handler(CommandHandler("help", cmd_help))
    application.add_handler(CommandHandler("profile", cmd_profile))
    application.add_handler(CommandHandler("addcoins", cmd_add_coins))  # для тестов
    application.add_handler(CommandHandler("openbox", cmd_openbox))
    application.add_handler(CommandHandler("cards", cmd_cards))
    application.add_handler(CommandHandler("inventory", cmd_inventory))
    application.add_handler(CommandHandler("mini", cmd_mini))
    application.add_handler(CommandHandler("reset", cmd_reset))

    # callback-кнопки
    application.add_handler(CallbackQueryHandler(cb_buy_box, pattern=r"^buy_box:"))
    application.add_handler(CallbackQueryHandler(cb_open_card, pattern=r"^open_card:"))
    application.add_handler(CallbackQueryHandler(cb_use_item, pattern=r"^use_item:"))

    # только polling — никакого webhook и Updater
    application.run_polling(
        allowed_updates=Update.ALL_TYPES,
        drop_pending_updates=True,
    )


if __name__ == "__main__":
    main()
